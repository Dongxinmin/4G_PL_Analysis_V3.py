import datetime
import pandas as pd
import openpyxl
import numpy as np
import time
#from graphviz import Digraph  # 画图用库
# 载入数据
df = pd.read_excel('CSSR Analysis UMTS.xlsx')
the_day = datetime.date.today()-datetime.timedelta(days=1)
the_day1 = datetime.date.today()-datetime.timedelta(days=2)
the_day2 = datetime.date.today()-datetime.timedelta(days=3)
print(the_day)

#df.loc[:, '900946:3G Call Setup Success Rate (CS)_MTN(%)_0218'] = df['900946:3G Call Setup Success Rate (CS)_MTN(%)_0218'].str.replace("%", "").astype('str')
#df.loc[:, '900947:Call Setup Success Rate (PS)_MTN(%)0218'] = df['900947:Call Setup Success Rate (PS)_MTN(%)0218'].str.replace("%", "").astype('str')
#--获取昨天的数据
df['csfails']=df['901057:Nr of RRC Connection Fails (CS)_MTN']+df['901060:Nr of RAB Establishment Fails (CS)_MTN']
df['psfails']=df['901083:Nr of RRC Connection Fails (PS)_MTN']+df['901122:Nr of RAB Establishment Fails (PS)_MTN']
df['CS_CSSR']=np.array(df['900946:3G Call Setup Success Rate (CS)_MTN(%)_0218'])*np.array(100)
df['PS_CSSR']=np.array(df['900947:Call Setup Success Rate (PS)_MTN(%)0218'])*np.array(100)
df['csrrcfails']=df['901057:Nr of RRC Connection Fails (CS)_MTN']
df['csrabfails']=df['901060:Nr of RAB Establishment Fails (CS)_MTN']
df['psrrcfails']=df['901083:Nr of RRC Connection Fails (PS)_MTN']
df['psrabfails']=df['901122:Nr of RAB Establishment Fails (PS)_MTN']
df['date']=df['Start Time'].str[0:10]
df2=df[['date','RNC Managed NE Name','NodeB Name','Cell Name','CS_CSSR','csfails','csrrcfails','csrabfails','PS_CSSR','psfails','psrrcfails','psrabfails']]
df2.apply(pd.to_numeric, errors='coerce')
df2[df2.date == str(the_day)].to_excel('3G Top cells.xlsx',index=False,sheet_name="data_log")
#----------

#----------
aa=pd.read_excel('3G Top cells.xlsx',sheet_name='data_log')
aa['NW_Cell_CS_Contri'] = (100*aa['csfails']/aa['csfails'].sum()).round(2)
aa['NW_Cell_CS_RANK'] = aa['csfails'].rank(ascending=False, method='max')
aa['NW_Cell_PS_Contri'] = (100*aa['psfails']/aa['psfails'].sum()).round(2)
aa['NW_Cell_PS_RANK'] = aa['psfails'].rank(ascending=False, method='max')
aa.to_excel('3G Top cells.xlsx',index=False,sheet_name='data_log')
#---

# 获得每天NW RNC的汇总
res_file = '3G Top cells.xlsx'
writer = pd.ExcelWriter(res_file,mode='a',index=False,engine='openpyxl')

summarys1= df.groupby(['date']).sum()
summarys1['3G_CS_RRCSR']=np.array(summarys1['901044:Nr of RRC Connection Successes (CS)_MTN']) / np.array(summarys1['901056:Nr of RRC Connection Requests (CS)_MTN']) *np.array(100)
summarys1['3G_CS_RABSR']=np.array(summarys1['901058:Nr of RAB Establishment Successes (CS)_MTN']) / np.array(summarys1['901059:Nr of RAB Establishment Attempts (CS)_MTN']) *np.array(100)
summarys1['3G_PS_RRCSR']=np.array(summarys1['901082:Nr of RRC Connection Successes (PS)_MTN']) / np.array(summarys1['901061:Nr of RRC Connection Requests (PS)_MTN']) *np.array(100)
summarys1['3G_PS_RABSR']=np.array(summarys1['901100:Nr of RAB Establishment Successes (PS)_MTN']) / np.array(summarys1['901101:Nr of RAB Establishment Attempts (PS)_MTN']) *np.array(100)
summarys1['3G_CS_CSSR'] =np.array(summarys1['3G_CS_RRCSR']) * np.array(summarys1['3G_CS_RABSR'])/np.array(100)
summarys1['3G_PS_CSSR'] =np.array(summarys1['3G_PS_RRCSR']) * np.array(summarys1['3G_PS_RABSR'])/np.array(100)
summarys1['3G_CS_CSSR_PCT_CHANGE'] =summarys1['3G_CS_CSSR'].pct_change().round(5)
summarys1['3G_PS_CSSR_PCT_CHANGE'] =summarys1['3G_PS_CSSR'].pct_change().round(5)
summarys2 =pd.DataFrame(summarys1[['3G_CS_CSSR','3G_CS_RRCSR','3G_CS_RABSR','3G_PS_CSSR','3G_PS_RRCSR','3G_PS_RABSR','3G_CS_CSSR_PCT_CHANGE','3G_PS_CSSR_PCT_CHANGE']])
summarys2.to_excel(writer, index=True, sheet_name="3G_Summary")

summarys3= df.groupby(['RNC Managed NE Name','date']).sum()
summarys3['RNC_CS_RRCSR']=np.array(summarys3['901044:Nr of RRC Connection Successes (CS)_MTN']) / np.array(summarys3['901056:Nr of RRC Connection Requests (CS)_MTN']) *np.array(100)
summarys3['RNC_CS_RABSR']=np.array(summarys3['901058:Nr of RAB Establishment Successes (CS)_MTN']) / np.array(summarys3['901059:Nr of RAB Establishment Attempts (CS)_MTN']) *np.array(100)
summarys3['RNC_PS_RRCSR']=np.array(summarys3['901082:Nr of RRC Connection Successes (PS)_MTN']) / np.array(summarys3['901061:Nr of RRC Connection Requests (PS)_MTN']) *np.array(100)
summarys3['RNC_PS_RABSR']=np.array(summarys3['901100:Nr of RAB Establishment Successes (PS)_MTN']) / np.array(summarys3['901101:Nr of RAB Establishment Attempts (PS)_MTN']) *np.array(100)
summarys3['RNC_CS_CSSR'] =np.array(summarys3['RNC_CS_RRCSR']) * np.array(summarys3['RNC_CS_RABSR'])/np.array(100)
summarys3['RNC_PS_CSSR'] =np.array(summarys3['RNC_PS_RRCSR']) * np.array(summarys3['RNC_PS_RABSR'])/np.array(100)
summarys3['RNC_CS_CSSR_PCT_CHANGE'] =summarys3['RNC_CS_CSSR'].pct_change().round(5)
summarys3['RNC_PS_CSSR_PCT_CHANGE'] =summarys3['RNC_PS_CSSR'].pct_change().round(5)
summarys4 =pd.DataFrame(summarys3[['RNC_CS_CSSR','RNC_CS_RRCSR','RNC_CS_RABSR','RNC_PS_CSSR','RNC_PS_RRCSR','RNC_PS_RABSR','RNC_CS_CSSR_PCT_CHANGE','RNC_PS_CSSR_PCT_CHANGE']])
summarys4.to_excel(writer, index=True, sheet_name="RNC_Summary")

summarys5= df.groupby(['NodeB Name','date']).sum()
summarys5['SITE_CS_RRCSR']=np.array(summarys5['901044:Nr of RRC Connection Successes (CS)_MTN']) / np.array(summarys5['901056:Nr of RRC Connection Requests (CS)_MTN']) *np.array(100)
summarys5['SITE_CS_RABSR']=np.array(summarys5['901058:Nr of RAB Establishment Successes (CS)_MTN']) / np.array(summarys5['901059:Nr of RAB Establishment Attempts (CS)_MTN']) *np.array(100)
summarys5['SITE_PS_RRCSR']=np.array(summarys5['901082:Nr of RRC Connection Successes (PS)_MTN']) / np.array(summarys5['901061:Nr of RRC Connection Requests (PS)_MTN']) *np.array(100)
summarys5['SITE_PS_RABSR']=np.array(summarys5['901100:Nr of RAB Establishment Successes (PS)_MTN']) / np.array(summarys5['901101:Nr of RAB Establishment Attempts (PS)_MTN']) *np.array(100)
summarys5['SITE_CS_CSSR'] =np.array(summarys5['SITE_CS_RRCSR']) * np.array(summarys5['SITE_CS_RABSR'])/np.array(100)
summarys5['SITE_PS_CSSR'] =np.array(summarys5['SITE_PS_RRCSR']) * np.array(summarys5['SITE_PS_RABSR'])/np.array(100)
summarys5['SITE_csrrcfails']=summarys5['901057:Nr of RRC Connection Fails (CS)_MTN']
summarys5['SITE_csrabfails']=summarys5['901060:Nr of RAB Establishment Fails (CS)_MTN']
summarys5['SITE_psrrcfails']=summarys5['901083:Nr of RRC Connection Fails (PS)_MTN']
summarys5['SITE_psrabfails']=summarys5['901122:Nr of RAB Establishment Fails (PS)_MTN']
summarys5['SITE_CS_Fail_Number'] = summarys5['901057:Nr of RRC Connection Fails (CS)_MTN']+summarys5['901060:Nr of RAB Establishment Fails (CS)_MTN']
summarys5['SITE_PS_Fail_Number'] = summarys5['901083:Nr of RRC Connection Fails (PS)_MTN']+summarys5['901122:Nr of RAB Establishment Fails (PS)_MTN']
summarys5['SITE_CS_CSSR_PCT_CHANGE'] =summarys5['SITE_CS_CSSR'].pct_change().round(5)
summarys5['SITE_PS_CSSR_PCT_CHANGE'] =summarys5['SITE_PS_CSSR'].pct_change().round(5)
summarys5['SITE_CS_Fail_PCT_CHANGE'] =summarys5['SITE_CS_Fail_Number'].pct_change().round(5)
summarys5['SITE_PS_Fail_PCT_CHANGE'] =summarys5['SITE_PS_Fail_Number'].pct_change().round(5)
summarys6 =pd.DataFrame(summarys5[['SITE_CS_CSSR','SITE_CS_RRCSR','SITE_CS_RABSR','SITE_PS_CSSR','SITE_PS_RRCSR','SITE_PS_RABSR','SITE_csrrcfails','SITE_csrabfails','SITE_psrrcfails','SITE_psrabfails','SITE_CS_Fail_Number','SITE_PS_Fail_Number','SITE_CS_CSSR_PCT_CHANGE','SITE_PS_CSSR_PCT_CHANGE','SITE_CS_Fail_PCT_CHANGE','SITE_PS_Fail_PCT_CHANGE']])
summarys6.to_excel(writer, index=True, sheet_name="Site_Summary")

writer.save()

#------RNC的数据画图
from openpyxl.chart import Series,LineChart, Reference
wb=openpyxl.load_workbook('3G Top cells.xlsx')
sheets = wb.sheetnames
print(sheets)

sheet = wb[sheets[1]]
maxRow = sheet.max_row  # 行
maxColumn = sheet.max_column   #列
chart = LineChart()  # 图表对象
chart.title = "3G KPI"
for j in range(2, maxColumn-1):
    data = Reference(sheet,min_col=j,min_row= 1,max_col=j,max_row=8)  # 涉及数据
    print(data)
    seriesObj = Series(data, title_from_data= True)  # 创建series对象
    chart.append(seriesObj)  # 添加到chart中
sheet.add_chart(chart, "K2")  # 将图表添加到 sheet中

sheet = wb[sheets[2]]
maxRow = sheet.max_row  # 行
maxColumn = sheet.max_column   #列
chart = LineChart()  # 图表对象
chart.title = "BZV_RNC301"
for j in range(3, maxColumn-1):
    data = Reference(sheet,min_col=j,min_row= 1,max_col=j,max_row=8)  # 涉及数据
    print(data)
    seriesObj = Series(data, title_from_data= True)  # 创建series对象
    chart.append(seriesObj)  # 添加到chart中
sheet.add_chart(chart, "K2")  # 将图表添加到 sheet中
titles = ['RNC_CS_CSSR','RNC_CS_RRCSR','RNC_CS_RABSR','RNC_PS_CSSR','RNC_PS_RRCSR','RNC_PS_RABSR']
chart2 = LineChart()  # 图表对象
chart2.title = "BZV_RNC302"
for j in range(3, maxColumn-1):
    data = Reference(sheet,min_col=j,min_row= 9,max_col=j,max_row=15)  # 涉及数据
    print(data)
    seriesObj = Series(data,title= titles[(j-3)])  # 创建series对象
    chart2.append(seriesObj)  # 添加到chart中
sheet.add_chart(chart2, "K16")  # 将图表添加到 sheet中
chart3 = LineChart()  # 图表对象
chart3.title = "BZV_RNC305"
for j in range(3, maxColumn-1):
    data = Reference(sheet,min_col=j,min_row= 16,max_col=j,max_row=22)  # 涉及数据
    print(data)
    seriesObj = Series(data,title= titles[(j-3)])  # 创建series对象
    chart3.append(seriesObj)  # 添加到chart中
sheet.add_chart(chart3, "K30")  # 将图表添加到 sheet中

chart4 = LineChart()  # 图表对象
chart4.title = "PNR_RNC303"
for j in range(3, maxColumn-1):
    data = Reference(sheet,min_col=j,min_row= 23,max_col=j,max_row=29)  # 涉及数据
    print(data)
    seriesObj = Series(data,title= titles[(j-3)])  # 创建series对象
    chart4.append(seriesObj)  # 添加到chart中
sheet.add_chart(chart4, "S2")  # 将图表添加到 sheet中

chart5 = LineChart()  # 图表对象
chart5.title = "PNR_RNC304"
for j in range(3, maxColumn-1):
    data = Reference(sheet,min_col=j,min_row= 30,max_col=j,max_row=36)  # 涉及数据
    print(data)
    seriesObj = Series(data,title= titles[(j-3)])  # 创建series对象
    chart5.append(seriesObj)  # 添加到chart中
sheet.add_chart(chart5, "S16")  # 将图表添加到 sheet中

chart6 = LineChart()  # 图表对象
chart6.title = "PNR_RNC306"
for j in range(3, maxColumn-1):
    data = Reference(sheet,min_col=j,min_row= 37,max_col=j,max_row=43)  # 涉及数据
    print(data)
    seriesObj = Series(data,title= titles[(j-3)])  # 创建series对象
    chart6.append(seriesObj)  # 添加到chart中
sheet.add_chart(chart6, "S30")  # 将图表添加到 sheet中

wb.save("3G Top cells " + str(time.strftime("%Y-%m-%d %H-%M-%S")) + ".xlsx")

